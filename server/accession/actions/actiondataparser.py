# -*- coding: utf-8; -*-
#
# @file parsedata
# @brief Parsing of CSV and XLSX tables of input for actions
# @author Frédéric SCHERMA (INRA UMR1095)
# @date 2018-03-06
# @copyright Copyright (c) 2018 INRA/CIRAD
# @license MIT (see LICENSE file)
# @details 

from django.utils.translation import ugettext_lazy as _
from openpyxl import Workbook, load_workbook

from accession.actions.actionstepformat import ActionStepFormat


class ActionDataParser(object):

    def __init__(self):
        self._data = []
        self._columns = []

    @property
    def data(self):
        return self._data

    @property
    def columns(self):
        return self._columns

    def parse_csv(self, buffer):
        header = buffer.readline().decode('utf-8')

        # detect separator
        if ',' in header:
            separator = ','
        elif ';' in header:
            separator = ';'
        elif '\t' in header:
            separator = '\t'
        else:
            raise ImportError(_("Unsupported CSV format"))

        columns = header.split(separator)
        num_cols = len(columns)

        if num_cols <= 0 or num_cols > 100:
            raise ImportError(_("Number of columns must be comprised between 1 to 100"))

        # map to interest of the column
        for col in columns:
            if col == "accession_id":
                self._columns.append(ActionStepFormat.IO_ACCESSION_ID)
            elif col == "batch_id":
                self._columns.append(ActionStepFormat.IO_BATCH_ID)
            elif col == "descriptor_id":
                self._columns.append(ActionStepFormat.IO_DESCRIPTOR)
            else:
                self._columns.append(-1)

        row = 0

        # content
        for line in buffer:
            content = line.decode('utf-8').rstrip('\n').split(separator)
            if len(content) != num_cols:
                raise ImportError(_("Invalid CSV row %i") % (row+1))

            self._data.append(content)

            row += 1

        return len(self._data)

    def parse_xlsx(self, buffer):
        wb = load_workbook(buffer)

        # shn = wb.get_sheet_names()
        ws = wb.active
        num_cols = 101

        columns = []

        # count filled cells for the first row
        for i in range(1, 101):
            cell = ws.cell(row=1, column=i)
            if not cell.value:
                num_cols = i - 1
                break

            columns.append(cell.value)

        if num_cols <= 0 or num_cols > 100:
            raise ImportError(_("Number of columns must be comprised between 1 to 100"))

        # map to interest of the column
        for col in columns:
            if col == "accession_id":
                self._columns.append(ActionStepFormat.IO_ACCESSION_ID)
            elif col == "batch_id":
                self._columns.append(ActionStepFormat.IO_BATCH_ID)
            elif col == "descriptor_id":
                self._columns.append(ActionStepFormat.IO_DESCRIPTOR)
            else:
                self._columns.append(-1)

        for row in ws.iter_rows(min_row=2, max_col=num_cols, max_row=100000):
            empty = 0
            row_data = []

            for cell in row:
                if not cell.value:
                    empty += 1

                row_data.append(cell.value)

            # empty row, mean end
            if empty == num_cols:
                break

            self._data.append(row_data)

        return len(self._data)
