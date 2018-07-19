# -*- coding: utf-8; -*-
#
# @file actiondataexporter
# @brief collgate 
# @author Frédéric SCHERMA (INRA UMR1095)
# @date 2018-03-08
# @copyright Copyright (c) 2018 INRA/CIRAD
# @license MIT (see LICENSE file)
# @details

import io

from django.utils.translation import ugettext_lazy as _
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from accession.actions.actionstepformat import ActionStepFormat
from accession.models import Accession, Batch, ActionData, ActionDataType
from descriptor.models import Descriptor


class ActionDataExporter(object):

    def __init__(self, action_controller, step_index):
        self._action_controller = action_controller
        self._step_index = step_index
        self._columns = []

        self._size = 0
        self._mime_type = ""
        self._file_ext = ""

        data_format = self._action_controller.get_step_data_format(self._step_index)
        for f in data_format:
            if f == ActionStepFormat.IO_ACCESSION_ID:
                self._columns.append('accession_name')
                self._columns.append('accession_id')
            elif f == ActionStepFormat.IO_BATCH_ID:
                self._columns.append('batch_name')
                self._columns.append('batch_id')
            elif f == ActionStepFormat.IO_DESCRIPTOR:
                self._columns.append('descriptor_xxx')
            else:
                pass

    def export_data_as_csv(self):
        data = self._action_controller.get_step_data(self._step_index)
        data_format = self._action_controller.get_step_data_format(self._step_index)

        output = io.BytesIO()
        header = ','.join(self._columns) + '\n'
        output.write(header.encode('utf-8'))

        cols = [[] for x in range(0, len(data_format))]

        if len(data_format) == 1:
            # single column is direct
            if data_format[0] == ActionStepFormat.IO_ACCESSION_ID:
                cols[0] = Accession.objects.filter(id__in=data).iterator()
            elif data_format[0] == ActionStepFormat.IO_BATCH_ID:
                cols[0] = Batch.objects.filter(id__in=data).iterator()
            elif data_format[0] == ActionStepFormat.IO_DESCRIPTOR:
                cols[0] = Descriptor.objects.filter(id__in=data).iterator()
            else:
                cols[0] = None
        else:
            # multiple columns, sub-array
            for f in data_format:
                if f == ActionStepFormat.IO_ACCESSION_ID:
                    cols[f] = Accession.objects.filter(id__in=data[f]).iterator()
                elif f == ActionStepFormat.IO_BATCH_ID:
                    cols[f] = Batch.objects.filter(id__in=data[f]).iterator()
                elif f == ActionStepFormat.IO_DESCRIPTOR:
                    cols[f] = Descriptor.objects.filter(id__in=data[f]).iterator()
                else:
                    cols[f] = None

        num_rows = len(data)
        for x in range(0, num_rows):
            row_content = []

            for f in data_format:
                if f == ActionStepFormat.IO_ACCESSION_ID:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                elif f == ActionStepFormat.IO_BATCH_ID:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                elif f == ActionStepFormat.IO_DESCRIPTOR:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                else:
                    row_content += [""]

                output.write((','.join(row_content) + '\n').encode('utf-8'))

        self._size = output.tell()
        output.seek(0, io.SEEK_SET)

        self._mime_type = 'text/csv'
        self._file_ext = ".csv"

        return output

    def export_data_as_xslx(self):
        data = self._action_controller.get_step_data(self._step_index)
        data_format = self._action_controller.get_step_data_format(self._step_index)

        output = io.BytesIO()

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        ws.append(self._columns)

        cols = [[] for x in range(0, len(data_format))]

        if len(data_format) == 1:
            if data_format[0] == ActionStepFormat.IO_ACCESSION_ID:
                cols[0] = Accession.objects.filter(id__in=data).iterator()
            elif data_format[0] == ActionStepFormat.IO_BATCH_ID:
                cols[0] = Batch.objects.filter(id__in=data).iterator()
            elif data_format[0] == ActionStepFormat.IO_DESCRIPTOR:
                cols[0] = Descriptor.objects.filter(id__in=data).iterator()
            else:
                cols[0] = None
        else:
            for f in data_format:
                if f == ActionStepFormat.IO_ACCESSION_ID:
                    cols[f] = Accession.objects.filter(id__in=data[f]).iterator()
                elif f == ActionStepFormat.IO_BATCH_ID:
                    cols[f] = Batch.objects.filter(id__in=data[f]).iterator()
                elif f == ActionStepFormat.IO_DESCRIPTOR:
                    cols[f] = Descriptor.objects.filter(id__in=data[f]).iterator()
                else:
                    cols[f] = None

        num_rows = len(data)
        for x in range(0, num_rows):
            row_content = []

            for f in data_format:
                if f == ActionStepFormat.IO_ACCESSION_ID:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                elif f == ActionStepFormat.IO_BATCH_ID:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                elif f == ActionStepFormat.IO_DESCRIPTOR:
                    entity = next(cols[f])
                    row_content += [entity.name, str(entity.id)]
                else:
                    row_content += [""]

                ws.append(row_content)

        output.write(save_virtual_workbook(wb))

        self._size = output.tell()
        output.seek(0, io.SEEK_SET)

        self._mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        self._file_ext = ".xlsx"

        return output

    @property
    def size(self):
        return self._size

    @property
    def mime_type(self):
        return self._mime_type

    @property
    def file_ext(self):
        return self._file_ext
